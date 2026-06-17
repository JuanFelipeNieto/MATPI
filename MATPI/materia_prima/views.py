from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.core.paginator import Paginator
from django.db import transaction
from django.utils import timezone
from .models import MateriaPrima, Lote
from usuarios.models import Administrador
import openpyxl
from datetime import datetime
import re
from django.views.decorators.http import require_http_methods, require_GET

MSG_NO_PERMISOS = "No tienes permisos para realizar esta acción."
MSG_DATOS_INVALIDOS = "Los datos no son válidos"

# Función auxiliar para validar si el ID en sesión es Administrador
def check_admin(request):
    id_sesion = request.session.get('usuario_id')
    return Administrador.objects.filter(usuario_id=id_sesion).exists()

# Función auxiliar para obtener un lote por ID
def get_lote_or_404(id_lote):
    return get_object_or_404(Lote, pk=id_lote)

# Función auxiliar para obtener materia prima por ID del POST
def get_materia_prima_from_post(post_data):
    id_materia = post_data.get('txt_id')
    return get_object_or_404(MateriaPrima, pk=id_materia)

# Función auxiliar para extraer datos de materia prima del POST
def extract_materia_prima_data(post_data):
    return {
        'nombre': post_data.get('txt_nombre'),
        'unidad': post_data.get('txt_unidad'),
        'cantidad_unidad': int(post_data.get('txt_cantidad_unidad', 1)),
        'tipo': post_data.get('txt_tipo', 'Comida'),
        'cantidad': int(post_data.get('txt_cantidad', 0)),
        'f_ingreso': post_data.get('txt_fecha_ingreso')
    }

# Función auxiliar para desempacar datos de materia prima
def unpack_materia_prima_data(data):
    return (
        data['nombre'],
        data['unidad'],
        data['cantidad_unidad'],
        data['tipo'],
        data['cantidad'],
        data['f_ingreso']
    )

# --- VISTA PRINCIPAL (LISTADO) ---

@require_GET
def listar_materia_prima(request):
    id_sesion = request.session.get('usuario_id')
    if not id_sesion:
        return redirect('login')

    es_admin = check_admin(request)
    query = request.GET.get('buscar', '')

    if query:
        materia_primas = MateriaPrima.objects.filter(nombre_materia_prima__icontains=query).order_by('nombre_materia_prima')
    else:
        materia_primas = MateriaPrima.objects.all().order_by('nombre_materia_prima')

    paginator = Paginator(materia_primas, 10)
    page_number = request.GET.get('page')
    materia_primas_paginated = paginator.get_page(page_number)

    return render(request, 'materia_prima/listar.html', {
        'materia_primas': materia_primas_paginated,
        'es_admin': es_admin,
        'buscar': query
    })

# --- GESTIÓN DE MATERIA PRIMA (CREACIÓN ABIERTA A CAJERO Y ADMIN) ---

@require_GET
def mostrar_registro_materia_prima(request):
    id_sesion = request.session.get('usuario_id')
    if not id_sesion:
        return redirect('login')

    # Solo el administrador puede registrar materia prima
    es_admin = check_admin(request)
    if not es_admin:
        messages.error(request, "Solo el administrador puede registrar materia prima.")
        return redirect('listar_materia_prima')

    return render(request, 'materia_prima/registrar.html', {
        'es_admin': es_admin,
        'fecha_actual': timezone.now()
    })

@require_http_methods(["GET", "POST"])
def registrar_materia_prima(request):
    # Verificamos que sea administrador
    if not check_admin(request):
        messages.error(request, MSG_NO_PERMISOS)
        return redirect('listar_materia_prima')

    if request.method == 'POST':
        data = extract_materia_prima_data(request.POST)
        nombre, unidad, cantidad_unidad, tipo, cantidad, f_ingreso = unpack_materia_prima_data(data)

        try:
            with transaction.atomic():
                materia = MateriaPrima.objects.create(
                    nombre_materia_prima=nombre,
                    unidad_medida=unidad,
                    cantidad_por_unidad=cantidad_unidad,
                    tipo=tipo,
                )

                if cantidad > 0:
                    Lote.objects.create(
                        materia_prima=materia,
                        cantidad_inicial=cantidad,
                        cantidad_actual=cantidad,
                        fecha_ingreso=f_ingreso or timezone.now(),
                        fecha_vencimiento=None
                    )

                messages.success(request, f"Materia prima '{nombre}' registrada exitosamente.")
        except Exception as e:
            messages.error(request, f"Error al registrar: {str(e)}")

        return redirect('listar_materia_prima')
    return redirect('mostrar_registro_materia_prima')

@require_GET
def pre_editar_materia_prima(request, id):
    es_admin = check_admin(request)
    # Aquí sí mantenemos el bloqueo de seguridad
    if not es_admin:
        messages.error(request, "Acceso denegado. Solo el administrador puede editar registros.")
        return redirect('listar_materia_prima')

    materia_prima = get_object_or_404(MateriaPrima, pk=id)
    return render(request, 'materia_prima/editar.html', {
        'materia_prima': materia_prima,
        'es_admin': es_admin
    })

@require_http_methods(["GET", "POST"])
def editar_materia_prima(request):
    if not check_admin(request):
        messages.error(request, MSG_NO_PERMISOS)
        return redirect('listar_materia_prima')

    if request.method == 'POST':
        materia = get_materia_prima_from_post(request.POST)

        materia.nombre_materia_prima = request.POST.get('txt_nombre')
        materia.unidad_medida        = request.POST.get('txt_unidad')
        materia.cantidad_por_unidad  = int(request.POST.get('txt_cantidad_unidad', 1))
        materia.tipo                 = request.POST.get('txt_tipo', 'Comida')
        materia.save()

        messages.success(request, "Información de la materia prima actualizada correctamente.")

    return redirect('listar_materia_prima')

@require_http_methods(["GET", "POST"])
def eliminar_materia_prima(request, id):
    if not check_admin(request):
        messages.error(request, "No tienes permisos para eliminar materia prima.")
        return redirect('listar_materia_prima')

    materia = get_object_or_404(MateriaPrima, pk=id)

    # Identificar productos afectados antes de eliminar
    productos_afectados = [detalle.producto for detalle in materia.detalles_producto.all()]

    materia.delete()

    # Recalcular cada producto afectado para actualizar stock y descripción automática
    from productos.views import recalcular_stock_producto
    for producto in productos_afectados:
        recalcular_stock_producto(producto)

    messages.success(request, "Materia prima eliminada correctamente y productos actualizados.")
    return redirect('listar_materia_prima')


@require_GET
def ver_lotes(request, id_materia):
    id_sesion = request.session.get('usuario_id')
    if not id_sesion: return redirect('login')

    es_admin = check_admin(request)
    materia = get_object_or_404(MateriaPrima, pk=id_materia)
    # Lotes con stock disponible primero, luego por vencimiento
    lotes = materia.lotes.filter(cantidad_actual__gt=0).order_by('fecha_vencimiento')
    lotes_agotados = materia.lotes.filter(cantidad_actual=0).order_by('-fecha_ingreso')[:10]

    return render(request, 'materia_prima/lotes.html', {
        'materia': materia,
        'lotes': lotes,
        'lotes_agotados': lotes_agotados,
        'es_admin': es_admin,
        'fecha_actual': timezone.now()
    })

@require_GET
def pre_editar_lote(request, id_lote):
    if not check_admin(request):
        messages.error(request, "Acceso denegado. Solo administradores pueden editar lotes.")
        return redirect('listar_materia_prima')

    lote = get_lote_or_404(id_lote)
    return render(request, 'materia_prima/editar_lote.html', {
        'lote': lote,
        'cantidad_actual_str': "{:.2f}".format(lote.cantidad_actual),
        'es_admin': True
    })

@require_http_methods(["GET", "POST"])
def editar_lote(request):
    if not check_admin(request):
        messages.error(request, "Permiso denegado.")
        return redirect('listar_materia_prima')

    if request.method == 'POST':
        id_lote = request.POST.get('txt_id')
        lote = get_lote_or_404(id_lote)

        from decimal import Decimal
        try:
            lote.cantidad_actual = Decimal(request.POST.get('txt_cantidad', 0))
        except (ValueError, TypeError):
            lote.cantidad_actual = Decimal(0)

        lote.fecha_vencimiento = request.POST.get('txt_fecha_vencimiento') or None
        lote.save()

        messages.success(request, f"Lote #{lote.id} actualizado correctamente.")
        return redirect('ver_lotes', id_materia=lote.materia_prima.id)

    return redirect('listar_materia_prima')

@require_http_methods(["GET", "POST"])
def eliminar_lote(request, id_lote):
    if not check_admin(request):
        messages.error(request, "Permiso denegado.")
        return redirect('listar_materia_prima')

    lote = get_lote_or_404(id_lote)
    id_materia = lote.materia_prima.id
    lote.delete()

    messages.success(request, "Lote eliminado correctamente.")
    return redirect('ver_lotes', id_materia=id_materia)

def _validar_fila_materia_prima(row):
    if len(row) < 4:
        raise ValueError(MSG_DATOS_INVALIDOS)
    nombre, unidad, cant_unidad, tipo = row[:4]
    if not nombre:
        return None
    try:
        cant_unidad = int(cant_unidad or 1)
    except (ValueError, TypeError):
        raise ValueError(MSG_DATOS_INVALIDOS)
    unidad = unidad or 'und'
    tipo = tipo or 'Comida'
    return {
        'nombre_materia_prima': nombre,
        'unidad_medida': unidad,
        'cantidad_por_unidad': cant_unidad,
        'tipo': tipo
    }

def _verificar_duplicado_materia_prima(datos):
    return MateriaPrima.objects.filter(
        nombre_materia_prima__iexact=datos['nombre_materia_prima'],
        unidad_medida__iexact=datos['unidad_medida'],
        cantidad_por_unidad=datos['cantidad_por_unidad'],
        tipo=datos['tipo']
    ).exists()

@require_http_methods(["GET", "POST"])
def importar_materia_prima_excel(request):
    if not check_admin(request):
        messages.error(request, MSG_NO_PERMISOS)
        return redirect('listar_materia_prima')

    if request.method != 'POST' or not request.FILES.get('archivo_excel'):
        return render(request, 'materia_prima/importar_materia_prima.html')

    archivo = request.FILES['archivo_excel']
    if not archivo.name.endswith('.xlsx'):
        messages.error(request, "Solo se permiten archivos .xlsx")
        return redirect('importar_materia_prima_excel')

    try:
        wb = openpyxl.load_workbook(archivo)
        sheet = wb.active
        rows = list(sheet.iter_rows(min_row=2, values_only=True))

        materias_a_crear = []
        for row in rows:
            datos = _validar_fila_materia_prima(row)
            if not datos:
                continue
            if _verificar_duplicado_materia_prima(datos):
                messages.error(request, "No se puede importar el archivo debido a que los datos ya existen o ya fueron ingresados")
                return redirect('listar_materia_prima')
            materias_a_crear.append(datos)

        creados = 0
        if materias_a_crear:
            with transaction.atomic():
                for datos in materias_a_crear:
                    MateriaPrima.objects.create(**datos)
                    creados += 1

        if creados > 0:
            messages.success(request, f"Se importaron {creados} materias primas correctamente.")

    except ValueError as e:
        messages.error(request, str(e))
        return redirect('importar_materia_prima_excel')
    except Exception as e:
        messages.error(request, f"Error al procesar el archivo: {str(e)}")

    return redirect('listar_materia_prima')


def _buscar_materia_prima_por_nombre(nombre_completo):
    nombre_completo = str(nombre_completo).strip()
    materia = MateriaPrima.objects.filter(nombre_materia_prima__iexact=nombre_completo).first()
    if not materia:
        if match := re.match(r"^(.*?)\s*\(([\d.,]+)\s+(.*)\)$", nombre_completo):
            nombre_base = match.group(1).strip()
            equiv_str = match.group(2).replace(',', '.')
            unidad_str = match.group(3).strip()
            try:
                equiv_int = int(float(equiv_str))
                materia = MateriaPrima.objects.filter(
                    nombre_materia_prima__iexact=nombre_base,
                    cantidad_por_unidad=equiv_int,
                    unidad_medida__iexact=unidad_str
                ).first()
            except Exception:
                pass
    return materia

def _validar_fila_lote(row, row_idx):
    if len(row) < 2:
        raise ValueError(MSG_DATOS_INVALIDOS)
    nombre_completo, cantidad, f_vencimiento, precio = row[:4]
    if not nombre_completo or cantidad is None:
        return None
    try:
        cantidad = float(cantidad)
        if precio is not None:
            precio = float(precio)
    except (ValueError, TypeError):
        raise ValueError(MSG_DATOS_INVALIDOS)

    materia = _buscar_materia_prima_por_nombre(nombre_completo)
    if not materia:
        raise ValueError(f"Error en fila {row_idx}: Materia prima '{nombre_completo}' no encontrada.")

    if isinstance(f_vencimiento, str):
        try:
            f_vencimiento = datetime.strptime(f_vencimiento, '%Y-%m-%d').date()
        except ValueError:
            f_vencimiento = None
    elif isinstance(f_vencimiento, datetime):
        f_vencimiento = f_vencimiento.date()

    return {
        'materia_prima': materia,
        'cantidad_inicial': cantidad,
        'fecha_vencimiento': f_vencimiento,
        'precio_unidad': precio
    }

@require_http_methods(["GET", "POST"])
def importar_lotes_excel(request):
    if not check_admin(request):
        messages.error(request, MSG_NO_PERMISOS)
        return redirect('listar_materia_prima')

    if request.method != 'POST' or not request.FILES.get('archivo_excel'):
        return render(request, 'materia_prima/importar_lotes.html')

    archivo = request.FILES['archivo_excel']
    if not archivo.name.endswith('.xlsx'):
        messages.error(request, "Solo se permiten archivos .xlsx")
        return redirect('importar_lotes_excel')

    try:
        wb = openpyxl.load_workbook(archivo)
        sheet = wb.active
        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        lotes_a_crear = []

        for row_idx, row in enumerate(rows, start=2):
            datos = _validar_fila_lote(row, row_idx)
            if not datos:
                continue
            if Lote.objects.filter(
                materia_prima=datos['materia_prima'],
                cantidad_inicial=datos['cantidad_inicial'],
                fecha_vencimiento=datos['fecha_vencimiento'],
                precio_unidad=datos['precio_unidad']
            ).exists():
                messages.error(request, "el lote ya existe")
                return redirect('listar_materia_prima')
            lotes_a_crear.append(datos)

        if lotes_a_crear:
            with transaction.atomic():
                for data in lotes_a_crear:
                    Lote.objects.create(
                        materia_prima=data['materia_prima'],
                        cantidad_inicial=data['cantidad_inicial'],
                        cantidad_actual=data['cantidad_inicial'],
                        fecha_ingreso=timezone.now(),
                        fecha_vencimiento=data['fecha_vencimiento'],
                        precio_unidad=data['precio_unidad']
                    )
            messages.success(request, f"Se importaron {len(lotes_a_crear)} lotes correctamente.")

    except ValueError as e:
        messages.error(request, str(e))
        return redirect('importar_lotes_excel')
    except Exception as e:
        messages.error(request, f"Error al procesar el archivo: {str(e)}")

    return redirect('listar_materia_prima')
